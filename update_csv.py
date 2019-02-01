import  os
from openpyxl import load_workbook
filepath = r'C:\Users\sunil2\Desktop'
filename = 'Data for Sunil.xlsx'
file_abs_name = os.path.join(filepath, filename)
wb  = load_workbook(file_abs_name)
ws1 = wb['Data']
ws2 = wb['Ref']
data_mapper = {}

# Construct a mapping from refernece sheet
ws2_iterator = ws2.iter_rows()
next(ws2_iterator)
for row in ws2_iterator:
    data_mapper.setdefault(row[0].value, row[1].value)
#print(data_mapper)

ws1_iterator = ws1.iter_rows()
# Skip first row
next(ws1_iterator)
for row in ws1_iterator:
    row[0].value = data_mapper[row[0].value]

#Save as new updated_workbook    
wb.save('updated_workbook.xlsx')
